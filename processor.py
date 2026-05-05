import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def load_data(transactions_path: str = "transactions.xlsx",
              companies_path:    str = "companies.xlsx") -> tuple[pd.DataFrame, pd.DataFrame]:
    print("[processor] Loading data files …")
    transactions = pd.read_excel(transactions_path, dtype={"nip": str})
    companies    = pd.read_excel(companies_path,    dtype={"nip": str})
    print(f"[processor] transactions: {len(transactions)} rows | companies: {len(companies)} rows")
    return transactions, companies



def clean_transactions(df: pd.DataFrame) -> pd.DataFrame:

    df = df.copy()
    df["nip"] = df["nip"].astype(str).str.strip()

    before = len(df)
    df = df[df["nip"].str.fullmatch(r"\d{10}")]   
    df = df.dropna(subset=["price"])
    after = len(df)

    removed = before - after
    if removed:
        print(f"[processor] Removed {removed} invalid/incomplete rows during cleaning")

    return df



def aggregate_by_nip(df: pd.DataFrame) -> pd.DataFrame:
    aggregated = (
        df.groupby("nip", as_index=False)["price"]
        .sum()
        .rename(columns={"price": "total_value"})
    )
    aggregated["total_value"] = aggregated["total_value"].round(2)
    print(f"[processor] Aggregated to {len(aggregated)} unique NIPs")
    return aggregated



def enrich_with_company_names(aggregated: pd.DataFrame,
                               companies: pd.DataFrame) -> pd.DataFrame:
 
    companies_clean = companies[["nip", "company_name"]].copy()
    companies_clean["nip"] = companies_clean["nip"].astype(str).str.strip()

    merged = aggregated.merge(companies_clean, on="nip", how="left")
    merged["company_name"] = merged["company_name"].fillna("Unknown Company")

    unknown_count = (merged["company_name"] == "Unknown Company").sum()
    print(f"[processor] {unknown_count} NIPs have no matching company name")

    return merged[["nip", "company_name", "total_value"]]



def get_top_customers(df: pd.DataFrame, top_n: int = 10) -> pd.DataFrame:
    """Sort by total_value descending and return the top N customers."""
    top = df.sort_values("total_value", ascending=False).head(top_n).reset_index(drop=True)
    top.index += 1   
    print(f"[processor] Top {top_n} customers selected")
    return top



def save_output(df: pd.DataFrame, output_path: str = "output.xlsx") -> None:
    """
    Write the report to Excel and apply professional formatting with openpyxl:
    header row, alternating row colours, column widths, and currency format.
    """
    df.to_excel(output_path, index=True, index_label="Rank", sheet_name="Top Customers")

    wb = load_workbook(output_path)
    ws = wb.active

    HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")   
    ALT_ROW_FILL = PatternFill("solid", fgColor="DEEAF1")  
    HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    BODY_FONT    = Font(name="Arial", size=10)
    CENTER       = Alignment(horizontal="center", vertical="center")
    LEFT         = Alignment(horizontal="left",   vertical="center")
    THIN_BORDER  = Border(
        bottom=Side(style="thin", color="B8CCE4"),
        top=Side(style="thin", color="B8CCE4"),
    )

    for cell in ws[1]:
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = CENTER

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = ALT_ROW_FILL if row_idx % 2 == 0 else PatternFill()
        for cell in row:
            cell.font      = BODY_FONT
            cell.fill      = fill
            cell.border    = THIN_BORDER
            cell.alignment = CENTER if cell.column == 1 else LEFT

    for cell in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        for c in cell:
            c.number_format = '#,##0.00 "PLN"'

    column_widths = {"A": 8, "B": 16, "C": 40, "D": 20}
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 18

    wb.save(output_path)
    print(f"[processor] Report saved → {output_path}")



def run_pipeline(transactions_path: str = "transactions.xlsx",
                 companies_path:    str = "companies.xlsx",
                 output_path:       str = "output.xlsx") -> None:
    """Execute the complete ETL pipeline end-to-end."""
    transactions, companies = load_data(transactions_path, companies_path)
    clean_tx    = clean_transactions(transactions)
    aggregated  = aggregate_by_nip(clean_tx)
    enriched    = enrich_with_company_names(aggregated, companies)
    top_df      = get_top_customers(enriched, top_n=10)
    save_output(top_df, output_path)
    return top_df



if __name__ == "__main__":
    run_pipeline()

